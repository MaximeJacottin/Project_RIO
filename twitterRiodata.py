import os
import re
import torch
import requests
import tensorflow as tf
import warnings
import tweepy as tw
import pandas as pd
import numpy as np
from transformers import BertTokenizer, TFBertForSequenceClassification
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from datetime import date
from openpyxl import load_workbook

warnings.filterwarnings("ignore")#runs two times to not have the warnings


#test
#Get current date
today = date.today()
day = today.strftime("%Y-%m-%d")


#Basic input to fill for the code to works
consumer_key = ""
consumer_secret_key = ""
access_token_key = ""
access_secret_token_key = ""
key_words = "Bitcoin"
date_until = day


#Workbook sheet path 
workbook_path = 'C:/Users/Maxime/Desktop/Code/Projet Python/Projet Rio data/Sentiment_Analysis.xlsx'#You need to put your working path here

#Authantification
auth = tw.OAuthHandler(consumer_key, consumer_secret_key)
auth.set_access_token(access_token_key, access_secret_token_key)
api = tw.API(auth, wait_on_rate_limit=True)


#Search with the Twitter API the tweets given a keyword and a date until you want the tweets
def search(key_words, date_until):

    query = tw.Cursor(api.search_tweets, q=key_words, lang="en", until=date_until).items(10)
    tweet = [[tweet.created_at, tweet.text] for tweet in query]
    tweet_df = pd.DataFrame(data=tweet, 
                    columns=['created_at', 'text'])
    return tweet_df


 #Clean DataFrame   
def clean_df(tweet_df):
    
    tweet_txt_df = []

    for tweet in tweet_df['text'].values:
        if type(tweet) == np.float64:
            return ""
        twt_txt = tweet.lower()
        twt_txt = re.sub("'", "", twt_txt)
        twt_txt = re.sub("@[A-Za-z0-9_]+","", twt_txt)
        twt_txt = re.sub("#[A-Za-z0-9_]+","", twt_txt)
        twt_txt = re.sub(r'http\S+', '', twt_txt)
        twt_txt = re.sub('[()!?]', ' ', twt_txt)
        twt_txt = re.sub('\[.*?\]',' ', twt_txt)
        twt_txt = re.sub("[^a-z0-9]"," ", twt_txt)
        twt_txt = twt_txt.split()
        #stopwords = ["for", "on", "an", "a", "of", "and", "in", "the", "to", "from", "rt"]
        stopwords = ["rt"]
        twt_txt = [w for w in twt_txt if not w in stopwords]
        twt_txt = " ".join(word for word in twt_txt)
        tweet_txt_df.append(twt_txt)
        tweet_df['Clean_text'] = pd.Series(tweet_txt_df)
        tweet_df['Clean_text'].drop_duplicates(inplace=True)
        clean_tweet_df = tweet_df[['created_at', 'Clean_text']]        

    return clean_tweet_df

#NLP Model Incrementation using Bert
tokeniser = AutoTokenizer.from_pretrained('nlptown/bert-base-multilingual-uncased-sentiment')
model = AutoModelForSequenceClassification.from_pretrained('nlptown/bert-base-multilingual-uncased-sentiment')

#Calculating the sentiment score 
def sentiment_score(text):

    token = tokeniser.encode(text, return_tensors='pt')
    result = model(token)
    return int(torch.argmax(result.logits))+1


#Groupby Date (Daily)
def groupby_date_daily(df):

    df = df.groupby(pd.Grouper(key='created_at', freq='B')).mean()
    df['Date'] = df.index
    df['Date'] = df['Date'].apply(lambda a: pd.to_datetime(a).date())
    df.reset_index(inplace=True)
    df.dropna(inplace=True)
    df['sentiment'] = df['sentiment'].astype(int)
    df = df[['Date', 'sentiment']]
    return df


def to_workbook_file(dataframe, workbook_path):

    book = load_workbook(workbook_path)#Where you want to put your data
    writer = pd.ExcelWriter(workbook_path, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = True,header= False)

    writer.save()


#Main function to call
def main():

    df_1 = search(key_words, date_until)
    df_2 = clean_df(df_1)
    df = pd.DataFrame(df_2)
    df['text'] = pd.Series([x for x in df['Clean_text'] if len(x)>=30])
    df = df.drop(columns='Clean_text')
    df = df.dropna(axis=0)
    df['sentiment'] = df['text'].apply(lambda x: sentiment_score(x[:512]))
    df = groupby_date_daily(df)
    to_excel_file = to_workbook_file(df, workbook_path)


if __name__ == "__main__":
    main()